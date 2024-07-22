from __future__ import print_function

import os
import math
import torch
import torch.optim as optim
import random
import numpy as np
import openpyxl
import pandas as pd
import pickle
import nibabel as nib
from skimage.metrics import structural_similarity as SSIM
from sklearn.metrics import mean_absolute_error as MAE
seed = 1024


def set_seeds():
    
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

set_seeds()


def create_dir(mypath):
    """Create a directory if it does not exist."""
    try:
        os.makedirs(mypath)
    except OSError as exc:
        if os.path.isdir(mypath):
            pass
        else:
            raise

class TwoCropTransform:
    def __init__(self, transform):
        self.transform = transform

    def __call__(self, brain_image, mask, slice_no, gmm_adr, image_slice):
        arg_dict = {"brain_image": brain_image,
                     "mask": mask,
                     "slice_no": slice_no, "gmm_adr": gmm_adr}  
        transformed_image = self.transform(arg_dict)
        return [image_slice, transformed_image]
    

class AverageMeter(object):
    """Computes and stores the average and current value"""
    def __init__(self):
        self.reset()

    def reset(self):
        self.val = 0
        self.avg = 0
        self.sum = 0
        self.count = 0

    def update_other_metrics(self, val, n = 1):
        self.val = val
        self.sum += val * n 
        self.count += n
        self.avg = self.sum / self.count
        
        
def convert_pickle_to_nifti(adr_in, adr_out, data_type):
    
    data = pd.read_excel(adr_in)
    pickled_adr_list = data["pickle_adr"].tolist()
    nifti_image1_adr_list = []
    nifti_image2_adr_list = []
    Affine = np.eye(4)
    create_dir("Dataset/Data_For_Loader/saved_data/" + data_type + "_nifti")
    
    
    for indi in range(0, len(pickled_adr_list)):
        print(indi)
        pickle_adr = pickled_adr_list[indi]
        
        nifti_adr_image1 = pickle_adr.replace(data_type , data_type + "_nifti").replace(".txt", "_image1.nii.gz")
        nifti_adr_image2 = pickle_adr.replace(data_type , data_type + "_nifti").replace(".txt", "_image2.nii.gz")
        
        nifti_image1_adr_list.append(nifti_adr_image1)
        nifti_image2_adr_list.append(nifti_adr_image2)
        
        with open(pickle_adr, "rb") as handle:
            Data = pickle.load(handle)
            handle.close()
            nib.save(nib.Nifti1Image(Data[0].numpy(), Affine), nifti_adr_image1)
            nib.save(nib.Nifti1Image(Data[2].numpy(), Affine), nifti_adr_image2)
            
    data["nifi_image1"] = nifti_image1_adr_list
    data["nifi_image2"] = nifti_image2_adr_list
    
    data.to_excel(adr_out)



def accuracy(output, target, topk=(1,)):
    """Computes the accuracy over the k top predictions for the specified values of k"""
    with torch.no_grad():
        maxk = max(topk)
        batch_size = target.size(0)

        _, pred = output.topk(maxk, 1, True, True)
        pred = pred.t()
        correct = pred.eq(target.view(1, -1).expand_as(pred))

        res = []
        for k in topk:
            correct_k = correct[:k].view(-1).float().sum(0, keepdim=True)
            res.append(correct_k.mul_(100.0 / batch_size))
        return res
    
    
def write_into_excel_file(filename, value):

    wb = openpyxl.load_workbook(filename=filename)
    sheet = wb[wb.sheetnames[0]]
    new_row = value
    sheet.append(new_row)
    wb.save(filename)
    
def make_excel_file(filename):
    
    wb = openpyxl.Workbook()
    wb.save(filename)


def adjust_learning_rate(args, optimizer, epoch):
    lr = args.learning_rate
    if args.cosine:
        eta_min = lr * (args.lr_decay_rate ** 3)
        lr = eta_min + (lr - eta_min) * (
                1 + math.cos(math.pi * epoch / args.epochs)) / 2
    else:
        steps = np.sum(epoch > np.asarray(args.lr_decay_epochs))
        if steps > 0:
            lr = lr * (args.lr_decay_rate ** steps)
    #print("Learning rate:" + str(lr))

    for param_group in optimizer.param_groups:
        param_group['lr'] = lr


def warmup_learning_rate(args, epoch, batch_id, total_batches, optimizer):
    if args.warm and epoch <= args.warm_epochs:
        p = (batch_id + (epoch - 1) * total_batches) / \
            (args.warm_epochs * total_batches)
        lr = args.warmup_from + p * (args.warmup_to - args.warmup_from)

        for param_group in optimizer.param_groups:
            param_group['lr'] = lr


def set_optimizer(opt, model, step_no):
    if (step_no == 1):
        optimizer = optim.Adam(model.parameters(), lr=opt.learning_rate_step1, eps = opt.eps)
    elif (step_no == 2):
        optimizer = optim.Adam(model.parameters(), lr=opt.learning_rate_step2, eps = opt.eps)
    return optimizer


def save_model(model, optimizer, opt, epoch, save_file):
    print('==> Saving...')
    state = {
        'opt': opt,
        'model': model.state_dict(),
        'optimizer': optimizer.state_dict(),
        'epoch': epoch,
    }
    torch.save(state, save_file)
    del state
    
    
def calculate_cross_brain_SSIM(image1, image2, index, mask, train_adr_data):
    
    slice_index = train_adr_data.iloc[[index]]["slice"].values[0]
    mask_slice = mask[:, :, slice_index]
    image1 = np.multiply(image1, mask_slice) 
    image1 = np.multiply(image2, mask_slice) 
    
    return SSIM(image1, image2)
    
def calculate_cross_brain_SSIM0(image1, image2, index, mask, train_adr_data):
    
    return 0
    
    
def calculate_cross_brain_MAE(image1, image2, index, mask, train_adr_data):
    
    slice_index = train_adr_data.iloc[[index]]["slice"].values[0]
    mask_slice = mask[:, :, slice_index]
    image1 = image1[mask_slice>0]
    image2 = image2[mask_slice>0]
    
    return MAE(image1, image2)
    

def calculate_metrics(images, indecies, mask, train_adr_data):
    
    sums_SSIM = []
    sums_MAE = []
    for indi in range(0, len(images)):
        for indj in range(indi + 1, len(images)):
            sum_SSIM = 0
            sum_MAE = 0
            for indk in range(0, len(images[0])):
                sum_SSIM += calculate_cross_brain_SSIM0(images[indi][indk], images[indj][indk]
                                                     , indecies[indk], mask, train_adr_data)
                sum_MAE += calculate_cross_brain_MAE(images[indi][indk], images[indj][indk]
                                                     , indecies[indk], mask, train_adr_data)
            sums_SSIM.append(sum_SSIM)
            sums_MAE.append(sum_MAE)
    return sums_SSIM, sums_MAE
    
    
def calculate_performance(images, scanner_no, indecies, mask, data_info):
    
    n = (int)(((scanner_no)*(scanner_no-1))/2)
    
    # detatch the images
    for i in range(0, len(images)):
        images[i] = images[i].to("cpu")
        images[i] = images[i].tolist()
        for j in range(0, len(images[i])):
            images[i][j] = np.squeeze(np.array(images[i][j]))
         
    # calculate the metrics
    sums_SSIM, sums_MAE = calculate_metrics(images, indecies, mask, data_info)
    return sums_SSIM, sums_MAE
