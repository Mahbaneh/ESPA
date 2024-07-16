'''
Created on Dec 29, 2023

@author: MAE82
'''
import os
import sys
import argparse
import time
import random
import numpy as np
import pandas as pd
import nibabel as nib
from os.path import exists
from torch.utils import data
import openpyxl
import math
import torch
import torch.backends.cudnn as cudnn
import torch.nn as nn
import tensorboard_logger as tb_logger
import torch.optim as optim
import gc 


# Setting GPUs
gpu_number = "2"
os.environ["CUDA_DEVICE_ORDER"] = "PCI_BUS_ID"
os.environ["CUDA_VISIBLE_DEVICES"] = gpu_number
device = torch.device("cuda" if (torch.cuda.is_available()) else "cpu")

def set_seeds(seed1):
    random.seed(seed1)
    np.random.seed(seed1)
    torch.manual_seed(seed1)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False
    
def set_seeds_torch(seed):
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

def parse_option(gpu_number):

    parser = argparse.ArgumentParser('argument for training')

    parser.add_argument('--calculate_performance_Flag', type=bool, default=False,
                        help='calculate performances')
                        
    parser.add_argument('--c_dim', type=int, default=5,
                        help='')
                        
    parser.add_argument('--res_blocks', type=int, default=9,
                        help='')
                        
    parser.add_argument('--nz', type=int, default=192,
                        help='noise dimension')
                        
    parser.add_argument('--data_dir', type=str, default = "./Dataset/ESPA_Res_Training/ErternalScannerTrainingMISPEL",
                        help='directory to augmented images')
                        
    parser.add_argument('--aug_data_folder', type=str, default = "Augmented_train_data",
                        help='folder name of augmented data')
                        
    parser.add_argument('--train_excel_adr', type=str, default = "./Dataset/ESPA_Res_Training/ErternalScannerTrainingMISPEL/OASIS_train.xlsx",
                        help='directory to the list of augmented images')
                        
    parser.add_argument('--performance_Freq', type=int, default=2,
                        help='frequency of calculating performance measures')
                        
    parser.add_argument('--data_frequency_step1', type=int, default=5,
                        help='frequency of seeing new set of data') 
                        
    parser.add_argument('--data_frequency_step2', type=int, default=14,
                        help='frequency of seeing new set of data') 

    parser.add_argument('--print_freq', type=int, default=2,
                        help='print frequency')
    
    parser.add_argument('--no_latent_embeddings', type=int, default=6,
                        help='number of components in th elatent embedding (L in paper).')
    
    parser.add_argument('--head', type=str, default="linear_decoder",
                        help='Type of head network which is the decoder in MISPEL.')
    
    parser.add_argument('--mask_adr', type=str, default = "Dataset/ESPA_Res_Training/cropped_JHU_MNI_SS_T1_Brain_Mask.nii.gz",
                        help='print frequency')
    
    parser.add_argument('--output_excel_name_step1_train', type=str, default = "",
                        help='name of excel file for step 1 train losses')
    
    parser.add_argument('--output_excel_name_step1_validation', type=str, default = "",
                        help='name of excel file for step 1 validation losses')
    
    parser.add_argument('--output_excel_name_step2_train', type=str, default = "",
                        help='name of excel file for step 2 train losses')
    
    parser.add_argument('--output_excel_name_step2_validatio', type=str, default = "",
                        help='name of excel file for step 2 validation losses')
    
    parser.add_argument('--padding_axial_dim', type=int, default=2,
                        help='print frequency')
    
    parser.add_argument('--no_axial_slices', type=int, default=152,
                        help='print frequency')
    
    parser.add_argument('--image_input_info', type=str, default="./Dataset/one_selected_scanner_data.xlsx",
                        help='this excel file should have the info of the images')

    parser.add_argument('--batch_size', type=int, default= 32,
                        help='batch_size')
    
    parser.add_argument('--num_workers', type=int, default= 0,
                        help='num of workers to use')
    
    parser.add_argument('--scanner_no', type=int, default= 4,
                        help='num of scanners')
    
    parser.add_argument('--epochs', type=int, default=1000,
                        help='number of training epochs')
    
    parser.add_argument('--epochs_step1', type=int, default=2,
                        help='number of training epochs in step1')
    
    parser.add_argument('--epochs_step2', type=int, default=2,
                        help='number of training epochs in step2')
    
    parser.add_argument('--save_freq_step1', type=int, default=50,
                        help='frequency for saving model for step1')
    
    parser.add_argument('--save_freq_step2', type=int, default=50,
                        help='frequency for saving model for step2')
    
    parser.add_argument('--train_valid_split', type=float, default=0.8,
                        help='train validation split ratio')

    parser.add_argument('--lambda1', type=float, default = 0.3,
                        help='Step1_Lrecon_coefficient')
    
    parser.add_argument('--lambda2', type=float, default = 1.0,
                        help='Step1_Lcoupling_coefficient')
    
    parser.add_argument('--lambda3', type=float, default = 1.0,
                        help='Step2_Lrecon_coefficient')
    
    parser.add_argument('--lambda4', type=float, default = 4.0,
                        help='Step2_Lharm_coefficient')
    
    parser.add_argument('--learning_rate_step1', type=float, default=0.0001,
                        help='learning rate')
                        
    parser.add_argument('--learning_rate_step2', type=float, default=0.001,
                        help='learning rate')
                        
    parser.add_argument('--eps', type=float, default=0.00000001,
                        help='learning rate')                      
      
    parser.add_argument('--lr_decay_epochs', type=str, default='700,800,900',
                        help='where to decay lr, can be a list')
    parser.add_argument('--lr_decay_rate', type=float, default=0.1,
                        help='decay rate for learning rate')
    parser.add_argument('--weight_decay', type=float, default=1e-4,
                        help='weight decay')
    parser.add_argument('--momentum', type=float, default=0.9,
                        help='momentum')

    # model dataset
    parser.add_argument('--model', type=str, default='UNET')
    parser.add_argument('--dataset', type=str, default='cifar10',
                        choices=['cifar10', 'cifar100', 'path'], help='dataset')
    parser.add_argument('--mean', type=str, help='mean of dataset in path in form of str tuple')
    parser.add_argument('--std', type=str, help='std of dataset in path in form of str tuple')
    parser.add_argument('--data_folder', type=str, default=None, help='path to custom dataset')
    parser.add_argument('--size', type=int, default=32, help='parameter for RandomResizedCrop')
    
    # method
    parser.add_argument('--method', type=str, default='SimCLR',
                        choices=['SupCon', 'SimCLR'], help='choose method')
    
    # temperature
    parser.add_argument('--temp', type=float, default=0.5,
                        help='temperature for loss function')

    # other setting
    parser.add_argument('--cosine', action='store_true',
                        help='using cosine annealing')
    parser.add_argument('--syncBN', action='store_true',
                        help='using synchronized batch normalization')
    parser.add_argument('--warm', action='store_true',
                        help='warm-up for large batch training')
    parser.add_argument('--trial', type=str, default='0',
                        help='id for recording multiple runs')
    
    opt = parser.parse_args()
    
    ############ Added by Mahbaneh  ################
    opt.cosine = True
    opt.syncBN = False
    ############ Added by Mahbaneh  ################
    
    # check if dataset is path that passed required arguments
    if opt.dataset == 'path':
        assert opt.data_folder is not None \
            and opt.mean is not None \
            and opt.std is not None

    # set the path according to the environment
    if opt.data_folder is None:
        opt.data_folder = './datasets/'
    opt.model_path = './save' + gpu_number + '/SupCon/{}_models'.format(opt.dataset)
    opt.tb_path = './save' + gpu_number + '/SupCon/{}_tensorboard'.format(opt.dataset)

    iterations = opt.lr_decay_epochs.split(',')
    opt.lr_decay_epochs = list([])
    for it in iterations:
        opt.lr_decay_epochs.append(int(it))

    opt.model_name = '{}_{}_{}_lr_{}_decay_{}_bsz_{}_temp_{}_trial_{}'.\
        format(opt.method, opt.dataset, opt.model, opt.learning_rate_step1,
               opt.weight_decay, opt.batch_size, opt.temp, opt.trial)

    if opt.cosine:
        opt.model_name = '{}_cosine'.format(opt.model_name)

    opt.save_folder = os.path.join(opt.model_path, opt.model_name, "training")
    opt.save_folder1 = os.path.join(opt.model_path, opt.model_name, "training_step1")
    opt.save_folder2 = os.path.join(opt.model_path, opt.model_name, "training_step2")
    
    if not os.path.isdir(opt.save_folder):
        os.makedirs(opt.save_folder)
        os.makedirs(opt.save_folder1)
        os.makedirs(opt.save_folder2)
        
    # Excel files for epochs
    opt.output_excel_name_step1_train = os.path.join(opt.save_folder1, "train_epoches.xlsx")
    opt.output_excel_name_step1_validation = os.path.join(opt.save_folder1, "validation_epoches.xlsx")
    opt.output_excel_name_step2_train = os.path.join(opt.save_folder2, "train_epoches.xlsx")
    opt.output_excel_name_step2_validation = os.path.join(opt.save_folder2, "validation_epoches.xlsx")
    filenames = [opt.output_excel_name_step1_train, opt.output_excel_name_step1_validation,
                 opt.output_excel_name_step2_train, opt.output_excel_name_step2_validation]
                 
    if os.path.exists(os.path.join("save" + str(gpu_number), "Results")) == False:
        os.mkdir(os.path.join("save" + str(gpu_number), "Results"))
    
    for filename in filenames:
        if (exists(filename) == False): 
            make_excel_file(filename)
            write_into_excel_file(filename, ["Epochs", "Loss", "Loss1", "Loss2"])
            
    opt.tb_folder = os.path.join(opt.tb_path, opt.model_name)      
    return opt


def normalizing_images(slice, scale):
        return  ((scale[1] - scale[0]) * ((slice - slice.min())/(0.0001 + slice.max() - slice.min()))) + scale[0]
        
def save_model(model, optimizer, opt, epoch, save_file):
    print('==> Saving...')
    state = {
        'opt': opt,
        'model': model.state_dict(),
        'optimizer': optimizer.state_dict(),
        'epoch': epoch,
    }
    torch.save(state, save_file)
    del state


def set_optimizer(opt, model, step_no):
    if (step_no == 1):
        optimizer = optim.Adam(model.parameters(), lr=opt.learning_rate_step1, eps = opt.eps)
    elif (step_no == 2):
        optimizer = optim.Adam(model.parameters(), lr=opt.learning_rate_step2, eps = opt.eps)
    return optimizer

def write_into_excel_file(filename, value):

    wb = openpyxl.load_workbook(filename=filename)
    sheet = wb[wb.sheetnames[0]]
    new_row = value
    sheet.append(new_row)
    wb.save(filename)

def change_adrs(adrs, substring, string):
    for i in range(0, adrs.shape[0]):
        for j in range(0, adrs.shape[1]):
            adrs[i][j] = adrs[i][j].replace(substring, string)
    return adrs

def read_mask(opt):
    mask = nib.load(opt.mask_adr).get_fdata()
    return mask

def make_excel_file(filename):
    wb = openpyxl.Workbook()
    wb.save(filename)
    

def set_loader(data, shuffle_flag, opt):
    
    dataset = HarmonizationDataSet(data, opt.scanner_no)
    data_loader = torch.utils.data.DataLoader(
        dataset, batch_size=opt.batch_size, shuffle = shuffle_flag,
        num_workers=opt.num_workers, pin_memory=True, sampler = None)

    return data_loader 


def set_model(opt):
    
    model = UnSupConMISPEL(name = opt.model, head = opt.head, 
                           latent_embedding_no = opt.no_latent_embeddings, scanner_no = opt.scanner_no).to(device)

    if torch.cuda.is_available():
        if torch.cuda.device_count() > 1:
            model.encoder = torch.nn.DataParallel(model.encoder)
        cudnn.benchmark = True

    return model


def set_losses(opt):
    # criterion step1
    criterion1 = MISPEL_losses(1, opt.lambda1, opt.lambda2, opt.lambda3, opt.lambda4, opt.batch_size, opt.scanner_no)
    
    # criterion step2
    criterion2 = MISPEL_losses(2, opt.lambda1, opt.lambda2, opt.lambda3, opt.lambda4, opt.batch_size, opt.scanner_no)

    if torch.cuda.is_available():
        criterion1 = criterion1.cuda()
        criterion2 = criterion2.cuda()

    return criterion1, criterion2


def train(train_loader, model, criterion, optimizer, epoch, opt, tarining_step_number):
    
    """one epoch training"""
    batch_time = AverageMeter()
    data_time = AverageMeter()

    end = time.time()
    train_loss = AverageMeter()
    train_loss1 = AverageMeter()
    train_loss2 = AverageMeter()
    
    model.train()
    torch.backends.cudnn.benchmark = True
    
    for idx, (images_list, _) in enumerate(train_loader):
        
        # redefine new images
        data_time.update_other_metrics(time.time() - end)
        images = [img.to(device) for img in images_list]
        
        # compute loss
        embeddings, reconstructed_images = model(images, tarining_step_number) 
        reconstructed_images = [torch.unsqueeze(reconstructed_image, dim = 1) for reconstructed_image in reconstructed_images]
        loss, loss1, loss2 = criterion(images, embeddings, reconstructed_images)

        # update metric
        train_loss.update_other_metrics(loss.item(), images[0].shape[0])
        train_loss1.update_other_metrics(loss1.item(), images[0].shape[0])
        train_loss2.update_other_metrics(loss2.item(), images[0].shape[0])

        # optimizer
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()

        # measure elapsed time
        batch_time.update_other_metrics(time.time() - end)
        end = time.time()

        # print info
        if (idx + 1) % opt.print_freq == 0:
            print('Train: [{0}][{1}/{2}]\t'
                  'BT {batch_time.val:.3f} ({batch_time.avg:.3f})\t'
                  'DT {data_time.val:.3f} ({data_time.avg:.3f})\t'
                  'loss {loss.val:.3f} ({loss.avg:.3f})'.format(
                   epoch, idx + 1, len(train_loader), batch_time=batch_time,
                   data_time=data_time, loss=train_loss))
            sys.stdout.flush()
    return train_loss.avg, train_loss1.avg, train_loss2.avg


def validation(valid_loader, model, criterion, optimizer, epoch, opt, 
               tarining_step_number):
    
    batch_time = AverageMeter()
    data_time = AverageMeter()
    validation_loss = AverageMeter()
    end = time.time()
    validation_loss1 = AverageMeter()
    validation_loss2 = AverageMeter()

    model.eval()
    with torch.no_grad():
        for idx, (images_list, _) in enumerate(valid_loader):
            data_time.update_other_metrics(time.time() - end)   
            
            # redefine new images
            images = [torch.tensor(img.detach().numpy()) for img in images_list]
            data_time.update_other_metrics(time.time() - end)
            images = [img.to(device) for img in images] 
                    
            embeddings, reconstructed_images = model(images, tarining_step_number) # tensor(btach-size * aug, features_dim)
            reconstructed_images = [torch.unsqueeze(reconstructed_image, dim = 1) for reconstructed_image in reconstructed_images]

            loss, loss1, loss2 = criterion(images, embeddings, reconstructed_images)            
            validation_loss.update_other_metrics(loss.item(), images[0].shape[0])
            validation_loss1.update_other_metrics(loss1.item(), images[0].shape[0])
            validation_loss2.update_other_metrics(loss2.item(), images[0].shape[0])
            
            # measure elapsed time
            batch_time.update_other_metrics(time.time() - end)
            end = time.time()
            
            if (idx + 1) % opt.print_freq == 0:
                print('Validation: [{0}][{1}/{2}]\t'
                  'BT {batch_time.val:.3f} ({batch_time.avg:.3f})\t'
                  'DT {data_time.val:.3f} ({data_time.avg:.3f})\t'
                  'loss {loss.val:.3f} ({loss.avg:.3f})'.format(
                   epoch, idx + 1, len(valid_loader), batch_time=batch_time,
                   data_time=data_time, loss=validation_loss))
                sys.stdout.flush()       
    return validation_loss.avg, validation_loss1.avg, validation_loss2.avg


def make_noise_mat(image_x, image_y, nz):
        noise = torch.randn(nz, dtype = torch.float)
        image_size = image_x * image_y
        image_mat_size = math.ceil(image_size/nz)
    
        repeated_noise = noise.repeat(image_mat_size)
        repeated_noise = torch.reshape(repeated_noise[0:image_size], (image_x, image_y))

        return repeated_noise[None, None, :, :]
    
    
def normalizing_images_gen(slice, scale):
    
    for indi in range(0, slice.shape[0]):
        slice[indi, :, :, :] = ((scale[indi][1] - scale[indi][0]) * ((slice[indi, :, :, :] - slice[indi, :, :, :].min())/(0.0001 + slice[indi, :, :, :].max() - slice[indi, :, :, :].min()))) + scale[indi][0]
    return slice
    
def cropping_images(data):
        
        x_ind = (int)((192 - 152)/ 2)
        y_ind = (int)((192 - 188)/ 2)
        data = data[:, :, x_ind: x_ind + 152, y_ind: y_ind + 188]
        return  data


def genearte_data(data, img_scale, model_path, opt, nz):

    # loading the generator 
    generator = GeneratorResNet(img_shape = img_shape, \
                                res_blocks = res_blocks, c_dim=c_dim).to(device)
    if (device.type == "cuda"):
        generator = nn.DataParallel(generator, [0])
    generator.load_state_dict(torch.load(model_path))

    # labels
    labels = [[0., 1., 0., 0., 0.],
              [0., 0., 1., 0., 0.],
              [0., 0., 0., 1., 0.],
              [0., 0., 0., 0., 1.]]
    target_scanner_lables = torch.tensor(labels, dtype = torch.float)
    data = torch.tensor(data, dtype = torch.float)
    conc_data = [torch.zeros((1, 1, 192, 192)) for indi in range(0, opt.scanner_no)]
    
    for indi in range(0, data.shape[0]):

        # make noise
        noise = make_noise_mat(data.shape[1], data.shape[2], nz).repeat((opt.scanner_no, 1, 1, 1)).to(device)
        slice = data[None, None, indi, :, :].repeat((opt.scanner_no, 1, 1, 1)).to(device)
        _, imgs = generator(slice, target_scanner_lables, noise)
        conc_data = [torch.cat((conc_data[indj], imgs[indj, None, :, :, :].clone().detach().cpu()), 0) for indj in range(0, opt.scanner_no)]
        
        del imgs
        del noise
        del slice
        gc.collect()
        torch.cuda.empty_cache()

    conc_data = [conc_data[indj][1:, :, :, :] for indj in range(0, opt.scanner_no)]  
    # Normalizing images
    conc_data = [normalizing_images_gen(conc_data[indj], img_scale) for indj in range(0, opt.scanner_no)]  
    # Resize images back to [1, 152, 188]
    conc_data = [cropping_images(conc_data[indj]) for indj in range(0, opt.scanner_no)] 
    return conc_data



def training(opt, optimizer, train_data, model, criterion, logger, 
             tarining_step_number, mask, nz):

    ############ set parameters for the appropriate step ################
    if tarining_step_number == 1:
        epochs = opt.epochs_step1
        save_folder = opt.save_folder1
        save_freq = opt.save_freq_step1
        excel_filename_train = opt.output_excel_name_step1_train
        excel_filename_validation = opt.output_excel_name_step1_validation
        step = "Step1"

    elif tarining_step_number == 2:
        epochs = opt.epochs_step2
        save_folder = opt.save_folder2
        save_freq = opt.save_freq_step2
        excel_filename_train = opt.output_excel_name_step2_train
        excel_filename_validation = opt.output_excel_name_step2_validation
        step = "Step2"
    else:
        print("Invalid step number!")
        
    # Refill train_dataloader
    # 1. generate data
    shuffle_flag = True
    # 2. define dataloader
    train_loader = set_loader(train_data, shuffle_flag, opt)

    for epoch in range(1, epochs + 1):
        
        ######################################################################################
        #################################### Train  #####################################
        

        loss, loss1, loss2 = train(train_loader, model, criterion, optimizer, epoch, opt,\
                                   tarining_step_number)
 
        # tensorboard logger
        logger.log_value(step + '_Train_loss', loss, epoch)
        logger.log_value(step + '_Train_loss1', loss1, epoch)
        logger.log_value(step + '_Train_loss2', loss2, epoch)
        logger.log_value(step + '_Train_learning_rate', optimizer.param_groups[0]['lr'], epoch)
        
        # write to excel file
        write_into_excel_file(excel_filename_train, [epoch, loss, loss1, loss2])

        # save model
        if epoch %save_freq == 0:
                save_file = os.path.join(save_folder, 'ckpt_epoch_{epoch}.pth'.format(epoch=epoch))
                save_model(model, optimizer, opt, epoch, save_file)



def two_step_training(opt, train_data, model, criterion1, criterion2, logger, 
                      mask, nz):

    ###################### Training: step1 ######################
    time1000 = time.time()
    training_step = 1
    optimizer_step1 = set_optimizer(opt, model, training_step)
    training(opt, optimizer_step1, train_data, model, criterion1,
              logger, training_step, mask, nz)
    time2000 = time.time()
    
    ###################### Training: step2 ######################
    training_step = 2
    # freezing the encoder  
    for param in model.encoders.parameters():
        param.requires_grad = False
    optimizer_step2 = set_optimizer(opt, model, training_step)
    training(opt, optimizer_step2, train_data, model, criterion2,
              logger, training_step, mask, nz)
    time3000 = time.time()

    print(" ")
    print("##################################### Summary ##############################################")
    print("step 1-start:" + str(time.asctime( time.localtime( time1000))))
    print("step 1-end:" + str(time.asctime( time.localtime( time2000))))
    print("step 2-end:" + str(time.asctime( time.localtime( time3000))))

    
class AverageMeter(object):
    """Computes and stores the average and current value"""
    def __init__(self):
        self.reset()

    def reset(self):
        self.val = 0
        self.avg = 0
        self.sum = 0
        self.count = 0
        
    def update_other_metrics(self, val, n = 1):
        self.val = val
        self.sum += val * n 
        self.count += n
        self.avg = self.sum / self.count
    
###########################################################################################################
#                                            Dataset
###########################################################################################################    
class HarmonizationDataSet(data.Dataset):   
    def __init__(self,
                 data,
                 scanner_no
                 ):
        self.data = data
        self.data = torch.tensor(self.data, dtype = torch.float)
        self.scanner_no = scanner_no
        
    def __len__(self):
        return (int)(self.data[0].shape[0])

    def __getitem__(self, index: int):
        #list of n images where n is the number of scanners
        images = [torch.unsqueeze(self.data[indi][index], 0) for indi in range(0, self.scanner_no)]
        return images, index
        
    
class DataGenerationDataSet(data.Dataset):   
    def __init__(self,
                 data_orig_scale, 
                 data,
                 scanne_no,
                 nz
                 ):
        self.data = data
        self.data_orig_scale = data_orig_scale
        self.data = torch.tensor(self.data, dtype = torch.float).to(device)
        self.labels = [[0., 1., 0., 0., 0.],
                       [0., 0., 1., 0., 0.],
                       [0., 0., 0., 1., 0.],
                       [0., 0., 0., 0., 1.]]
        self.target_scanner_lables = torch.tensor(self.labels, dtype = torch.float).to(device)
        self.scanne_no = scanne_no
        self.nz = nz
        self.scale = data_orig_scale

        
    def __len__(self):
        return (int)(self.data.shape[0])
       
    def normalizing_images(self, slice, scale):
        for indi in range(0, slice.shape[0]):
            slice[indi, :, :] = ((scale[1] - scale[0]) * ((slice[indi, :, :] - slice[indi, :, :].min())/(0.0001 + slice[indi, :, :].max() - slice[indi, :, :].min()))) + scale[0]
        return slice
    
    def cropping_images(self, data):
        
        x_ind = (int)((192 - 152)/ 2)
        y_ind = (int)((192 - 188)/ 2)
        data = data[:, :, x_ind: x_ind + 152, y_ind: y_ind + 188]
        return  data

    def make_noise_mat(self, image_x, image_y, nz):
        noise = torch.randn(nz, dtype = torch.float)
        image_size = image_x * image_y
        image_mat_size = math.ceil(image_size/nz)
    
        repeated_noise = noise.repeat(image_mat_size)
        repeated_noise = torch.reshape(repeated_noise[0:image_size], (image_x, image_y))

        return repeated_noise[None, None, :, :]

    def __getitem__(self, index: int):
        #list of n images where n is the number of scanners
        # image shape (axial slices): [1, 152, 188]

        # make noise
        noise = self.make_noise_mat(self.data.shape[1], self.data.shape[2], self.nz).to(device).repeat((self.scanne_no, 1, 1, 1))
        slice = self.data[None, index, :, :].repeat((self.scanne_no, 1, 1, 1))
        # generate images
        _, imgs = self.generator(slice.detach(), self.target_scanner_lables, noise)
        # scaling back images
        imgs = self.normalizing_images(imgs, self.scale[index])
        # resize images back to [1, 152, 188]
        imgs = self.cropping_images(imgs)
        
        images = [imgs[indi] for indi in range(0, imgs.shape[0])] 
        return images, index


        
###########################################################################################################
#                                            Networks
###########################################################################################################
class ResidualBlock(nn.Module):
    def __init__(self, in_features):
        super(ResidualBlock, self).__init__()

        conv_block = [
            nn.Conv2d(in_features, in_features, 3, stride=1, padding=1, bias=False),
            nn.InstanceNorm2d(in_features, affine=True, track_running_stats=True),
            nn.ReLU(inplace=True),
            nn.Conv2d(in_features, in_features, 3, stride=1, padding=1, bias=False),
            nn.InstanceNorm2d(in_features, affine=True, track_running_stats=True),
        ]

        self.conv_block = nn.Sequential(*conv_block)

    def forward(self, x):
        return x + self.conv_block(x)


class GeneratorResNet(nn.Module):
    def __init__(self, img_shape=(1, 192, 192), res_blocks=9, c_dim=5):
        super(GeneratorResNet, self).__init__()
        channels, img_size, _ = img_shape

        # Initial convolution block
        # 1 in channels + c_dim + 1 is for noise
        model = [
            nn.Conv2d(channels + c_dim + 1, 64, 7, stride=1, padding=3, bias=False), 
            nn.InstanceNorm2d(64, affine=True, track_running_stats=True),
            nn.ReLU(inplace=True),
        ]

        # Downsampling
        curr_dim = 64
        for _ in range(2):
            model += [
                nn.Conv2d(curr_dim, curr_dim * 2, 4, stride=2, padding=1, bias=False),
                nn.InstanceNorm2d(curr_dim * 2, affine=True, track_running_stats=True),
                nn.ReLU(inplace=True),
            ]
            curr_dim *= 2

        # Residual blocks
        for _ in range(res_blocks):
            model += [ResidualBlock(curr_dim)]

        # Upsampling
        for _ in range(2):
            model += [
                nn.ConvTranspose2d(curr_dim, curr_dim // 2, 4, stride=2, padding=1, bias=False),
                nn.InstanceNorm2d(curr_dim // 2, affine=True, track_running_stats=True),
                nn.ReLU(inplace=True),
            ]
            curr_dim = curr_dim // 2

        # Output layer
        model += [nn.Conv2d(curr_dim, channels, 7, stride=1, padding=3)]

        self.model = nn.Sequential(*model)

    def forward(self, x, c, noise):
        c = c.view(c.size(0), c.size(1), 1, 1)
        c = c.repeat(1, 1, x.size(2), x.size(3))
        img = torch.cat((x, c, noise), 1)
        residual = self.model(img) + x
        return residual, torch.tanh(residual)

class Linear(nn.Module):
    
    def __init__(self, embedding_length):
        
        super().__init__()
        self.embedding_length = embedding_length
        self.weights = nn.Parameter(torch.Tensor(1, self.embedding_length))
        nn.init.kaiming_uniform_(self.weights, a=math.sqrt(5))

    
    def forward(self, x):

        x = x *  self.weights[..., None, None]
        x = x.sum(dim = 1)
        return x

class Down_module(nn.Module):
    
    def __init__(self, in_channels, out_channels, kernel_size, padding):
        super(Down_module, self).__init__()
        self.in_channels = in_channels 
        self.out_channels = out_channels
        self.kernel_size = kernel_size
        self.padding = padding
        
        self.Conv1 = nn.Conv2d(self.in_channels, self.out_channels, kernel_size=self.kernel_size, padding= self.padding)
        self.batchN1 = nn.BatchNorm2d(self.out_channels)
        self.relu1 = nn.ReLU(inplace=True)
        self.Conv2 = nn.Conv2d(self.out_channels, self.out_channels, kernel_size=self.kernel_size, padding= self.padding)
        self.batchN2 = nn.BatchNorm2d(self.out_channels)
        self.relu2 = nn.ReLU(inplace=True)
        self.Maxpool = nn.MaxPool2d(2)


    def forward(self, x):
        
        x = self.Conv1(x)
        x =  self.batchN1(x)
        x = self.relu1(x)
        x = self.Conv2(x)
        x = self.batchN2(x)
        x = self.relu2(x)
        return x
    
class Up_module(nn.Module):
    
    def __init__(self, in_channels, out_channels, kernel_size, padding):
        super(Up_module, self).__init__()
        self.in_channels = in_channels 
        self.out_channels = out_channels
        self.kernel_size = kernel_size
        self.padding = padding
        
        self.Upsample1 = nn.Upsample(scale_factor=2)
        self.Conv1 = nn.Conv2d(self.in_channels, self.out_channels, kernel_size=self.kernel_size, padding= self.padding)
        
        self.Conv2 = nn.Conv2d(self.in_channels, self.out_channels, kernel_size=self.kernel_size, padding= self.padding)
        self.batchN2 = nn.BatchNorm2d(self.out_channels)
        self.relu2 = nn.ReLU(inplace=True)
        
        self.Conv3 = nn.Conv2d(self.out_channels, self.out_channels, kernel_size=self.kernel_size, padding= self.padding)
        self.batchN3 = nn.BatchNorm2d(self.out_channels)
        self.relu3 = nn.ReLU(inplace=True)



    def forward(self, x, concatenated_layer):
        
        x = self.Upsample1(x)
        x = self.Conv1(x)
        x = torch.cat((concatenated_layer, x), dim = 1)

        x = self.Conv2(x)
        x = self.batchN2(x)
        x = self.relu2(x)
        
        x = self.Conv3(x)
        x = self.batchN3(x)
        x = self.relu3(x)
        return x
    
class generate_embedding(nn.Module):
    
    def __init__(self, in_channels, out_channels, kernel_size, padding):
        super(generate_embedding, self).__init__()
        self.in_channels = in_channels 
        self.out_channels = out_channels
        self.kernel_size = kernel_size
        self.padding = padding

        self.conv1 = nn.Conv2d(in_channels  = self.in_channels, out_channels = self.out_channels, kernel_size= self.kernel_size, padding = self.padding)
        self.batch1 = nn.BatchNorm2d(self.out_channels)
        self.Relu1 = nn.ReLU(inplace=True)    
        
    def forward(self, x):
        
        x =  self.conv1(x)
        x = self.batch1(x)
        x = self.Relu1(x)
        return x
    
       
class UNet_2D(nn.Module):
    
    def __init__(self, n_channels):
        
        super(UNet_2D, self).__init__()
        
        self.n_channels = n_channels # put other variables in here. 
        self.Down1 = Down_module(1, 32, 3, 1)
        self.Down2 = Down_module(32, 64, 3, 1)
        self.Down3 = Down_module(64, 128, 3, 1)
    
        self.Maxpool1 = nn.MaxPool2d(2)
        self.Maxpool2 = nn.MaxPool2d(2)
        
        self.Up1 = Up_module(128, 64, 3, 1)
        self.Up2 = Up_module(64, 32, 3, 1)
        
        self.generate_embedding1 = generate_embedding(32, self.n_channels, 3, 1)
        
     
    def forward(self, x):
        # Down1
        x = self.Down1(x)
        layer1_for_concatenation = x
        x = self.Maxpool1(x)
        # Down2
        x = self.Down2(x)
        layer2_for_concatenation = x
        x = self.Maxpool2(x)
        # Embedding
        x = self.Down3(x)
        # Up1
        x = self.Up1(x, layer2_for_concatenation)
        # Up2
        x = self.Up2(x, layer1_for_concatenation)
        # generate embedding
        x = self.generate_embedding1(x)
        return x


def UNET(embedding_no, **kwargs):
    return UNet_2D(embedding_no, **kwargs)
model_dict = {
    'UNET': UNET,
}

        
class UnSupConMISPEL(nn.Module):
    """U-Net + liners decoder"""
    
    def __init__(self, name = 'UNET', head = 'linear_decoder', latent_embedding_no = 6, scanner_no = 4): 
        super(UnSupConMISPEL, self).__init__()

        model_fun = model_dict[name]
        self.scanner_no = scanner_no
        self.encoders = nn.ModuleList([model_fun(latent_embedding_no) for i in range(scanner_no)])
        self.linears_decoders = nn.ModuleList([Linear(latent_embedding_no) for i in range(scanner_no)])
        
            
    def forward(self, x, tarining_step_number):
        # I want the whole unit in MISPEL: (1) complete Unet, and (2) linear functions. 
        embeddings = []
        recontructed_images = []
        
        for i in range(self.scanner_no):
            embeddings.append(self.encoders[i](x[i]))
            recontructed_images.append(self.linears_decoders[i](embeddings[i]))

        return embeddings, recontructed_images 

###########################################################################################################
#                                            losses
###########################################################################################################    
class MISPEL_losses(nn.Module):
    
    def __init__(self, step_number, lambda1, lambda2, lambda3, lambda4, batch_size, image_number): #*** Mah
        
        super(MISPEL_losses, self).__init__()
        self.step_number = step_number
        self.lambda1 = lambda1
        self.lambda2 = lambda2
        self.lambda3 = lambda3
        self.lambda4 = lambda4
        self.batch_size = batch_size
        self.sum = 0
        self.MAE = nn.L1Loss()
        self.image_number = image_number #*** Mah

    
    def l_coulpling(self, embedding_sets):
        ''' Extracting loss for all images of batch '''
        
        # 
        for ind in range(0, len(embedding_sets)):
            embedding_sets[ind] = embedding_sets[ind].flatten().unsqueeze(dim = 1)

        concatenated_embeddings = torch.cat(embedding_sets, dim = 1)
        variances = concatenated_embeddings.var(dim = 1)
        l_coulpling = variances.mean()
        return l_coulpling
    
    
    def l_reconstruction(self, original_images, reconstructed_images):
        
        # Sigma_i=1:M(MAE(x_i^j, xBar_i^j)) 
        self.sum = 0.0
        for i in range(0, original_images.shape[0]):
            self.sum += self.MAE(original_images[i, ...], reconstructed_images[i, ...])
            
        # This is average loss for each batch. 
        return (self.sum)/(self.batch_size)

    
    def l_harmonization(self, reconstructed_image_set):
        
        self.sum = 0.0
        
        for i in range(0, len(reconstructed_image_set[0])):
            for j in range(0, self.image_number):
                for k in range(j + 1, self.image_number):
                    self.sum += self.MAE(reconstructed_image_set[j][i], reconstructed_image_set[k][i])
        permutation_no = 2/((self.image_number) * (self.image_number - 1))

        return self.sum/((permutation_no) * (self.batch_size))
    
    
    def forward(self, original_images, embeddings, reconstructed_images):
        
        original_images = torch.cat(original_images, dim=0) 
        reconstructed_images_concatenated = torch.cat(reconstructed_images, dim=0) 
            
        # compute l_reconstruction0
        l_reocn = self.l_reconstruction(original_images, reconstructed_images_concatenated)
            
        if self.step_number == 1:
            # compute l_coulpling
            l_coup = self.l_coulpling(embeddings) 
            loss_step1 = self.lambda1 * l_reocn + self.lambda2 * l_coup  
            return loss_step1, l_reocn, l_coup     

        if self.step_number == 2:
            # compute l_harmonization
            l_harm = self.l_harmonization(reconstructed_images)
            loss_step2 = self.lambda3 * l_reocn + self.lambda4 * l_harm
            return loss_step2, l_reocn, l_harm

        else: 
            print("Invalid number for taring steps!")
            return 0.0
        
def resize_images(slice):
    
    size_x = 192
    size_y = 192 
    desired_size_x = 152
    desired_size_y = 188 

    resized_slice = np.zeros([1, size_x, size_y])
    x_changes = (int) ((size_x - desired_size_x)/2)
    y_changes = (int) ((size_y - desired_size_y)/2)
    resized_slice[:, x_changes: x_changes + desired_size_x, y_changes: y_changes + desired_size_y] = slice
    return resized_slice


def read_data(dir, adr, folder_name):
    
    # read excel file
    dir = os.path.join(dir, folder_name, "nifti")
    scanners = ["ge", "philips", "prisma", "trio"]
    image_info = pd.read_excel(adr)
    subject_names = image_info["SBJ_name"].values.tolist()
    list_stacked_slices = []
     
    for inds in range(0, len(scanners)):
        
        # Read the first image
        image_adr = os.path.join(dir, scanners[inds], subject_names[0] + ".nii.gz")
        stacked_slices = nib.load(image_adr).get_fdata()[:, :, 2:-2]
        
        for indi in range(1, len(subject_names)):
            image_adr = os.path.join(dir, scanners[inds], subject_names[indi] + ".nii.gz")
            stacked_slices = np.concatenate([stacked_slices, nib.load(image_adr).get_fdata()[:, :, 2:-2]],\
                                        axis = 2)
        stacked_slices = np.swapaxes(stacked_slices, 1, 2)
        stacked_slices = np.swapaxes(stacked_slices, 0, 1)
        list_stacked_slices.append(stacked_slices)
    return list_stacked_slices

 
def extract_scale_of_images(data):
    
    scales = []
    for indi in range(0, data.shape[0]):
        Min = round(np.min(data[indi, :, :]), 4)
        Max = round(np.max(data[indi, :, :]), 4)
        scales.append([Min, Max])
    return scales


def main_for_new_loader_paired(gpu_number, img_shape):

    set_seeds_torch(1024)
    opt = parse_option(gpu_number) # Read arquments   
    mask = read_mask(opt) # Read mask 
    
    # Read data
    train_data = read_data(opt.data_dir, opt.train_excel_adr, opt.aug_data_folder)
    shuffle_flag = True
     
    model = set_model(opt) # build model and criterion
    criterion1, criterion2 = set_losses(opt) # losses
  
    logger = tb_logger.Logger(logdir=opt.tb_folder, flush_secs=2) # tensorboard
    two_step_training(opt, train_data, model, criterion1,\
                       criterion2, logger, mask, opt.nz) # training  
    
    
Start = time.time() 
img_shape = (1, 192, 192)
main_for_new_loader_paired(gpu_number, img_shape)
end = time.time()

print("finished the run!")
print("Strat:" + str(time.asctime(time.localtime(Start))))
print("End:" + str(time.asctime( time.localtime(end))))  




